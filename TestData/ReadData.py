import openpyxl
vk = openpyxl.load_workbook(r"C:\Users\david\PycharmProjects\DDT\daten.xlsx")
#sh = vk ["eins"]
#print(sh.max_row)


def fetch_number_of_rows(Sheetname):
    sh = vk[Sheetname]
    return sh.max_row

def fetch_cell_data(Sheetname, row, cell):
    sh = vk[Sheetname]
    cell = sh.cell(row, cell)
    return cell.value

print(fetch_number_of_rows("eins"))
print(fetch_cell_data("eins",2,2))